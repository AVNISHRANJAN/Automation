"""qa_exporter.py - Excel exporter for QA automation report.

The preferred path uses openpyxl when it is installed. A small stdlib-only XLSX
writer is kept as a fallback so report generation still works in bare Python
environments.
"""

import os
from datetime import datetime
from urllib.parse import urlparse
from xml.sax.saxutils import escape
from zipfile import ZIP_DEFLATED, ZipFile


TEST_CASE_HEADERS = [
    "Test Case ID",
    "Test Scenario",
    "Steps to Execute",
    "Expected Result",
    "Actual Result",
    "Status",
    "Screenshot",
]


def _report_path(report, out_dir: str) -> str:
    os.makedirs(out_dir, exist_ok=True)
    host = urlparse(report.url).netloc.replace(".", "_").replace(":", "_")
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return os.path.join(out_dir, f"qa_report_{host}_{ts}.xlsx")


def _test_case_rows(report) -> list[list[str]]:
    rows = [TEST_CASE_HEADERS]
    for tc in report.test_cases:
        rows.append(
            [
                tc.test_case_id,
                tc.test_scenario,
                tc.steps_to_execute,
                tc.expected_result,
                tc.actual_result,
                tc.status,
                tc.screenshot_reference,
            ]
        )
    return rows


def _summary_rows(report) -> list[list[str]]:
    return [
        ["URL", report.url],
        ["Final URL", report.final_url],
        ["Started At", report.started_at],
        ["Finished At", report.finished_at],
        ["Pages Tested", getattr(report, "pages_tested", "")],
        ["Selectors File", getattr(report, "selector_file_path", "")],
        ["Total Test Cases", report.total],
        ["Passed", report.passed],
        ["Failed", report.failed],
    ]


def export_qa_xlsx(report, out_dir: str = "reports") -> str:
    path = _report_path(report, out_dir)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except Exception:
        _export_qa_xlsx_stdlib(report, path)
        return path

    wb = Workbook()
    ws = wb.active
    ws.title = "QA Test Cases"

    for row in _test_case_rows(report):
        ws.append(row)

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for col, header in enumerate(TEST_CASE_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    status_col = 6  # "Status" is column 6
    for row in range(2, ws.max_row + 1):
        status_cell = ws.cell(row=row, column=status_col)
        status_value = (status_cell.value or "").strip().lower()
        status_cell.alignment = Alignment(horizontal="center", vertical="center")
        if status_value == "pass":
            status_cell.fill = pass_fill
        elif status_value == "fail":
            status_cell.fill = fail_fill

    widths = {
        1: 16,   # Test Case ID
        2: 46,   # Test Scenario
        3: 60,   # Steps to Execute
        4: 44,   # Expected Result
        5: 50,   # Actual Result
        6: 12,   # Status
        7: 60,   # Screenshot
    }
    for col, width in widths.items():
        ws.column_dimensions[chr(64 + col)].width = width

    for row in range(2, ws.max_row + 1):
        for col in [2, 3, 4, 5, 7]:
            ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical="top")

    summary = wb.create_sheet("Summary")
    for row in _summary_rows(report):
        summary.append(row)

    summary.column_dimensions["A"].width = 22
    summary.column_dimensions["B"].width = 80
    for row in range(1, summary.max_row + 1):
        summary.cell(row=row, column=1).font = Font(bold=True)

    wb.save(path)
    return path


def _col_name(index: int) -> str:
    name = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        name = chr(65 + remainder) + name
    return name


def _sheet_xml(rows: list[list[str]]) -> str:
    xml_rows = []
    for row_idx, row in enumerate(rows, start=1):
        cells = []
        for col_idx, value in enumerate(row, start=1):
            cell_ref = f"{_col_name(col_idx)}{row_idx}"
            safe_value = escape(str(value or ""))
            cells.append(f'<c r="{cell_ref}" t="inlineStr"><is><t>{safe_value}</t></is></c>')
        xml_rows.append(f'<row r="{row_idx}">{"".join(cells)}</row>')

    max_col = max((len(row) for row in rows), default=1)
    dimension = f"A1:{_col_name(max_col)}{max(len(rows), 1)}"
    return f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <dimension ref="{dimension}"/>
  <sheetViews><sheetView workbookViewId="0"/></sheetViews>
  <sheetFormatPr defaultRowHeight="15"/>
  <sheetData>{"".join(xml_rows)}</sheetData>
</worksheet>'''


def _export_qa_xlsx_stdlib(report, path: str) -> None:
    files = {
        "[Content_Types].xml": '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
  <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
  <Default Extension="xml" ContentType="application/xml"/>
  <Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>
  <Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>
  <Override PartName="/xl/worksheets/sheet2.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>
</Types>''',
        "_rels/.rels": '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>
</Relationships>''',
        "xl/workbook.xml": '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
  <sheets>
    <sheet name="QA Test Cases" sheetId="1" r:id="rId1"/>
    <sheet name="Summary" sheetId="2" r:id="rId2"/>
  </sheets>
</workbook>''',
        "xl/_rels/workbook.xml.rels": '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>
  <Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet2.xml"/>
</Relationships>''',
        "xl/worksheets/sheet1.xml": _sheet_xml(_test_case_rows(report)),
        "xl/worksheets/sheet2.xml": _sheet_xml(_summary_rows(report)),
    }

    with ZipFile(path, "w", ZIP_DEFLATED) as xlsx:
        for filename, content in files.items():
            xlsx.writestr(filename, content)
