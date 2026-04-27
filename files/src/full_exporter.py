"""
full_exporter.py — HTML + JSON + Excel export for the full test report
"""

import json
import os
from datetime import datetime
from urllib.parse import urlparse
from xml.sax.saxutils import escape
from zipfile import ZIP_DEFLATED, ZipFile
from src.test_runner import TestReport


# ── Excel helpers (shared column layout with qa_exporter) ─────────────────────
TEST_CASE_HEADERS = [
    "Test Case Id",
    "Module",
    "Test Case Description",
    "Expected Result",
    "Screenshot",
    "Result",
]

# Maps module attribute name → friendly display name
_MODULE_LABELS = {
    "form":        "Form",
    "nav":         "Navigation",
    "auth":        "Authentication",
    "session":     "Session",
    "search":      "Search",
    "payment":     "Payment",
    "files":       "File Upload/Download",
    "errors":      "Error Handling",
    "performance": "Performance",
    "security":    "Security",
    "compat":      "Compatibility",
}

_MODULE_ORDER = [
    "form", "nav", "auth", "session", "search",
    "payment", "files", "errors", "performance", "security", "compat",
]


def _safe_name(url: str) -> str:
    host = urlparse(url).netloc.replace(".", "_").replace(":", "_")
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"fullreport_{host}_{ts}"


def export_full_json(report: TestReport, out_dir: str = "reports") -> str:
    os.makedirs(out_dir, exist_ok=True)
    filename = _safe_name(report.url) + ".json"
    filepath = os.path.join(out_dir, filename)

    def serialize_results(results):
        out = []
        for r in results:
            d = {
                "test": r.test,
                "passed": r.passed,
                "detail": r.detail,
            }
            if hasattr(r, "category"):   d["category"] = r.category
            if hasattr(r, "severity"):   d["severity"] = r.severity
            if hasattr(r, "value"):      d["value"] = r.value
            if hasattr(r, "unit"):       d["unit"] = r.unit
            out.append(d)
        return out

    modules = {
        "form":        report.form,
        "nav":         report.nav,
        "auth":        report.auth,
        "session":     report.session,
        "search":      report.search,
        "payment":     report.payment,
        "files":       report.files,
        "errors":      report.errors,
        "performance": report.performance,
        "security":    report.security,
        "compat":      report.compat,
    }

    data = {
        "meta": {
            "url": report.url,
            "final_url": report.final_url,
            "page_title": report.page_title,
            "scanned_at": datetime.now().isoformat(),
            "scan_time_s": report.scan_time_s,
        },
        "summary": {
            "total_tests": report.total_tests,
            "total_passed": report.total_passed,
            "total_failed": report.total_failed,
            "score_percent": report.score,
        },
        "modules": {
            name: {
                "passed": len(mod.passed),
                "total": len(mod.results),
                "results": serialize_results(mod.results),
            }
            for name, mod in modules.items()
        },
    }

    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

    return filepath


# ── Excel export for --full-test ───────────────────────────────────────────────

def _full_test_case_rows(report: TestReport) -> list[list]:
    """Convert all TestResult objects into Excel rows."""
    rows = [TEST_CASE_HEADERS]
    counter = 1
    for attr in _MODULE_ORDER:
        module_obj = getattr(report, attr, None)
        if module_obj is None:
            continue
        module_label = _MODULE_LABELS.get(attr, attr.capitalize())
        for r in module_obj.results:
            tc_id = f"TC_{counter:03d}"
            counter += 1
            # Description = test name + detail
            description = r.test
            if getattr(r, "detail", ""):
                description += f"\n{r.detail}"
            # Expected result: inferred from test name
            expected = f"{r.test} should pass"
            # Screenshot: captured only in --qa-report mode
            result_status = "Pass" if r.passed else "Fail"
            screenshot = "" if r.passed else "N/A (— screenshots captured in --qa-report mode only)"
            rows.append([tc_id, module_label, description, expected, screenshot, result_status])
    return rows


def _full_summary_rows(report: TestReport) -> list[list]:
    return [
        ["URL", report.url],
        ["Final URL", report.final_url],
        ["Page Title", report.page_title],
        ["Scan Time (s)", report.scan_time_s],
        ["Total Test Cases", report.total_tests],
        ["Passed", report.total_passed],
        ["Failed", report.total_failed],
        ["Score (%)", report.score],
    ]


def _col_name(index: int) -> str:
    name = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        name = chr(65 + remainder) + name
    return name


def _sheet_xml(rows: list[list]) -> str:
    xml_rows = []
    for row_idx, row in enumerate(rows, start=1):
        cells = []
        for col_idx, value in enumerate(row, start=1):
            cell_ref = f"{_col_name(col_idx)}{row_idx}"
            safe_value = escape(str(value or ""))
            cells.append(f'<c r="{cell_ref}" t="inlineStr"><is><t>{safe_value}</t></is></c>')
        xml_rows.append(f'<row r="{row_idx}">{"" .join(cells)}</row>')
    max_col = max((len(row) for row in rows), default=1)
    dimension = f"A1:{_col_name(max_col)}{max(len(rows), 1)}"
    return f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <dimension ref="{dimension}"/>
  <sheetViews><sheetView workbookViewId="0"/></sheetViews>
  <sheetFormatPr defaultRowHeight="15"/>
  <sheetData>{"" .join(xml_rows)}</sheetData>
</worksheet>'''


def _export_full_xlsx_stdlib(report: TestReport, path: str) -> None:
    files = {
        "[Content_Types].xml": '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
  <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
  <Default Extension="xml" ContentType="application/xml"/>
  <Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>
  <Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>
  <Override PartName="/xl/worksheets/sheet2.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>
</Types>''',
        "_rels/.rels": '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>
</Relationships>''',
        "xl/workbook.xml": '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
  <sheets>
    <sheet name="QA Test Cases" sheetId="1" r:id="rId1"/>
    <sheet name="Summary" sheetId="2" r:id="rId2"/>
  </sheets>
</workbook>''',
        "xl/_rels/workbook.xml.rels": '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>
  <Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet2.xml"/>
</Relationships>''',
        "xl/worksheets/sheet1.xml": _sheet_xml(_full_test_case_rows(report)),
        "xl/worksheets/sheet2.xml": _sheet_xml(_full_summary_rows(report)),
    }
    with ZipFile(path, "w", ZIP_DEFLATED) as xlsx:
        for filename, content in files.items():
            xlsx.writestr(filename, content)


def export_full_xlsx(report: TestReport, out_dir: str = "reports") -> str:
    """Export the full-test report as an Excel file (.xlsx) with the standard
    six-column QA layout:
        Test Case Id | Module | Test Case Description | Expected Result | Screenshot | Result
    """
    os.makedirs(out_dir, exist_ok=True)
    host = urlparse(report.url).netloc.replace(".", "_").replace(":", "_")
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = os.path.join(out_dir, f"fullreport_{host}_{ts}.xlsx")

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except Exception:
        _export_full_xlsx_stdlib(report, path)
        return path

    wb = Workbook()
    ws = wb.active
    ws.title = "QA Test Cases"

    for row in _full_test_case_rows(report):
        ws.append(row)

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    pass_fill   = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Style header row
    for col, header in enumerate(TEST_CASE_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Colour Result column
    result_col = 6
    for row_idx in range(2, ws.max_row + 1):
        result_cell = ws.cell(row=row_idx, column=result_col)
        result_value = (result_cell.value or "").strip().lower()
        result_cell.alignment = Alignment(horizontal="center", vertical="center")
        if result_value == "pass":
            result_cell.fill = pass_fill
        elif result_value == "fail":
            result_cell.fill = fail_fill

    # Column widths
    widths = {1: 16, 2: 24, 3: 70, 4: 44, 5: 20, 6: 14}
    for col, width in widths.items():
        ws.column_dimensions[chr(64 + col)].width = width

    # Wrap text for description columns
    for row_idx in range(2, ws.max_row + 1):
        for col in [2, 3, 4]:
            ws.cell(row=row_idx, column=col).alignment = Alignment(wrap_text=True, vertical="top")

    # Summary sheet
    summary = wb.create_sheet("Summary")
    for row in _full_summary_rows(report):
        summary.append(row)
    summary.column_dimensions["A"].width = 22
    summary.column_dimensions["B"].width = 80
    for row_idx in range(1, summary.max_row + 1):
        summary.cell(row=row_idx, column=1).font = Font(bold=True)

    wb.save(path)
    return path


def export_full_html(report: TestReport, out_dir: str = "reports") -> str:
    os.makedirs(out_dir, exist_ok=True)
    filename = _safe_name(report.url) + ".html"
    filepath = os.path.join(out_dir, filename)

    score = report.score
    score_color = "#16a34a" if score >= 80 else "#ca8a04" if score >= 50 else "#dc2626"

    modules_data = [
        ("11", "📝 Form Testing",           report.form),
        ("12", "🧭 Navigation Testing",     report.nav),
        ("13", "🔐 Authentication Testing", report.auth),
        ("14", "🍪 Session Testing",        report.session),
        ("15", "🔍 Search Testing",         report.search),
        ("16", "💳 Payment Testing",        report.payment),
        ("17", "📁 File Upload/Download",   report.files),
        ("18", "❌ Error Handling",         report.errors),
        ("19", "⚡ Performance Testing",    report.performance),
        ("20", "🔒 Security Testing",       report.security),
        ("21", "🌐 Compatibility Testing",  report.compat),
    ]

    def sev_badge(result):
        if not hasattr(result, "severity"): return ""
        colors = {"critical": ("#7f1d1d","#fca5a5"), "high": ("#7c2d12","#fdba74"),
                  "medium": ("#713f12","#fde047"), "low": ("#1e3a5f","#93c5fd")}
        bg, fg = colors.get(result.severity, ("#374151","#d1d5db"))
        return f'<span style="background:{bg};color:{fg};padding:1px 6px;border-radius:4px;font-size:10px;font-weight:700;margin-left:6px">{result.severity.upper()}</span>'

    def result_row(r):
        icon  = "✓" if r.passed else "✗"
        color = "#16a34a" if r.passed else "#dc2626"
        bg    = "#f0fdf4" if r.passed else "#fff7f7"
        cat   = f'<span style="color:#6b7280;font-size:10px;margin-right:8px">[{getattr(r,"category","—")}]</span>'
        sev   = sev_badge(r) if not r.passed else ""
        return f'''<tr style="background:{bg}">
          <td style="padding:6px 10px;color:{color};font-weight:700;font-size:14px;width:24px">{icon}</td>
          <td style="padding:6px 10px">{cat}<strong style="font-size:12px">{r.test}</strong>{sev}</td>
          <td style="padding:6px 10px;color:#6b7280;font-size:11px">{r.detail[:120]}</td>
        </tr>'''

    modules_html = ""
    summary_rows = ""

    for num, title, mod in modules_data:
        if not mod.results:
            continue
        passed = len(mod.passed)
        total  = len(mod.results)
        ratio  = (passed / total) if total else 1
        bar_color = "#16a34a" if ratio >= 0.8 else "#ca8a04" if ratio >= 0.5 else "#dc2626"
        bar_width = round(ratio * 100)

        rows_html = "".join(result_row(r) for r in mod.results)
        modules_html += f"""
        <div class="module">
          <div class="module-header" onclick="toggle('{num}')">
            <span class="module-title">{num}. {title}</span>
            <span class="module-score" style="color:{bar_color}">{passed}/{total}</span>
            <div class="progress-bar"><div class="progress-fill" style="width:{bar_width}%;background:{bar_color}"></div></div>
            <span class="chevron" id="chev-{num}">▼</span>
          </div>
          <div class="module-body" id="body-{num}">
            <table><tbody>{rows_html}</tbody></table>
          </div>
        </div>"""

        summary_rows += f'<tr><td style="padding:6px 12px">{num}. {title}</td><td style="padding:6px 12px;color:{bar_color};font-weight:700">{passed}/{total}</td><td style="padding:6px 12px"><div style="background:#e5e7eb;border-radius:4px;height:8px;width:120px"><div style="background:{bar_color};border-radius:4px;height:8px;width:{bar_width}%"></div></div></td></tr>'

    # Critical/High failures
    crit_items = [r for r in report.all_results if not r.passed and hasattr(r, "severity") and r.severity == "critical"]
    high_items = [r for r in report.all_results if not r.passed and hasattr(r, "severity") and r.severity == "high"]

    alerts_html = ""
    if crit_items:
        crit_list = "".join(f"<li>⚠️ {r.test} — {r.detail[:80]}</li>" for r in crit_items)
        alerts_html += f'<div class="alert critical"><strong>🚨 {len(crit_items)} CRITICAL Issue(s)</strong><ul>{crit_list}</ul></div>'
    if high_items:
        high_list = "".join(f"<li>{r.test} — {r.detail[:80]}</li>" for r in high_items[:5])
        alerts_html += f'<div class="alert high"><strong>⚠️ {len(high_items)} HIGH Severity Issue(s)</strong><ul>{high_list}</ul></div>'

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Full Test Report — {report.page_title or report.url}</title>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif; background: #0f172a; color: #e2e8f0; min-height: 100vh; }}
  .hero {{ background: linear-gradient(135deg, #1e293b 0%, #0f172a 100%); padding: 2rem 2.5rem; border-bottom: 1px solid #1e293b; }}
  .hero h1 {{ font-size: 22px; font-weight: 700; color: #f1f5f9; }}
  .hero p  {{ font-size: 13px; color: #64748b; margin-top: 4px; }}
  .score-circle {{ display:inline-block; width:72px; height:72px; border-radius:50%; background: conic-gradient({score_color} {score}%, #1e293b {score}%); display:flex; align-items:center; justify-content:center; font-size:18px; font-weight:800; color:{score_color}; float:right; margin-top:-8px; position:relative; }}
  .score-inner {{ position:absolute; background:#0f172a; width:56px; height:56px; border-radius:50%; display:flex; align-items:center; justify-content:center; }}
  .metrics {{ display:grid; grid-template-columns:repeat(auto-fit,minmax(120px,1fr)); gap:12px; padding:1.5rem 2.5rem; background:#0f172a; }}
  .metric {{ background:#1e293b; border:1px solid #334155; border-radius:10px; padding:14px; }}
  .metric .lbl {{ font-size:10px; color:#64748b; text-transform:uppercase; letter-spacing:.06em; }}
  .metric .val {{ font-size:26px; font-weight:700; margin-top:4px; }}
  .content {{ padding:1.5rem 2.5rem; }}
  .alert {{ border-radius:8px; padding:12px 16px; margin-bottom:14px; font-size:13px; }}
  .alert.critical {{ background:#7f1d1d; border:1px solid #dc2626; }}
  .alert.high     {{ background:#431407; border:1px solid #ea580c; }}
  .alert ul {{ margin-top:8px; padding-left:20px; }}
  .alert li {{ margin-top:4px; color:#fca5a5; }}
  .module {{ background:#1e293b; border:1px solid #334155; border-radius:10px; margin-bottom:12px; overflow:hidden; }}
  .module-header {{ display:flex; align-items:center; gap:12px; padding:14px 16px; cursor:pointer; user-select:none; }}
  .module-header:hover {{ background:#263249; }}
  .module-title {{ font-weight:600; font-size:14px; flex:1; }}
  .module-score {{ font-size:13px; font-weight:700; width:40px; text-align:right; }}
  .progress-bar {{ flex:1; max-width:100px; background:#334155; border-radius:4px; height:6px; }}
  .progress-fill {{ height:6px; border-radius:4px; transition: width .3s; }}
  .chevron {{ font-size:12px; color:#64748b; transition:transform .2s; }}
  .module-body {{ overflow:hidden; max-height:0; transition:max-height .3s ease; }}
  .module-body.open {{ max-height:9999px; }}
  table {{ width:100%; border-collapse:collapse; }}
  td {{ border-top:1px solid #1e293b; vertical-align:top; }}
  .summary-table {{ width:100%; background:#1e293b; border:1px solid #334155; border-radius:10px; overflow:hidden; margin-top:1rem; }}
  .summary-table th {{ background:#0f172a; padding:8px 12px; text-align:left; font-size:11px; color:#64748b; text-transform:uppercase; }}
  .summary-table td {{ padding:8px 12px; border-top:1px solid #334155; font-size:13px; }}
</style>
</head>
<body>
<div class="hero">
  <div class="score-circle"><div class="score-inner">{score}%</div></div>
  <h1>🌐 WebScanner — Full Test Report</h1>
  <p>{report.url} &nbsp;·&nbsp; {datetime.now().strftime('%d %b %Y %H:%M')} &nbsp;·&nbsp; {report.scan_time_s}s &nbsp;·&nbsp; {report.total_passed}/{report.total_tests} tests passed</p>
</div>

<div class="metrics">
  <div class="metric"><div class="lbl">Score</div><div class="val" style="color:{score_color}">{score}%</div></div>
  <div class="metric"><div class="lbl">Passed</div><div class="val" style="color:#16a34a">{report.total_passed}</div></div>
  <div class="metric"><div class="lbl">Failed</div><div class="val" style="color:#dc2626">{report.total_failed}</div></div>
  <div class="metric"><div class="lbl">Critical</div><div class="val" style="color:#dc2626">{len(crit_items)}</div></div>
  <div class="metric"><div class="lbl">High</div><div class="val" style="color:#ea580c">{len(high_items)}</div></div>
  <div class="metric"><div class="lbl">Scan Time</div><div class="val" style="font-size:18px;color:#94a3b8">{report.scan_time_s}s</div></div>
</div>

<div class="content">
  {alerts_html}
  {modules_html}

  <h2 style="margin:2rem 0 1rem;font-size:15px;color:#94a3b8">Module Summary</h2>
  <table class="summary-table">
    <thead><tr><th>Module</th><th>Passed</th><th>Progress</th></tr></thead>
    <tbody>{summary_rows}</tbody>
  </table>
</div>

<script>
function toggle(id) {{
  const body = document.getElementById('body-' + id);
  const chev = document.getElementById('chev-' + id);
  body.classList.toggle('open');
  chev.style.transform = body.classList.contains('open') ? 'rotate(180deg)' : '';
}}
// Open first 3 modules by default
['11','12','13'].forEach(id => {{ const b = document.getElementById('body-'+id); if(b) b.classList.add('open'); }});
</script>
</body></html>"""

    with open(filepath, "w", encoding="utf-8") as f:
        f.write(html)

    return filepath
